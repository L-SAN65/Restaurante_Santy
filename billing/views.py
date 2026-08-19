from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.http import FileResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from core.models import Role

from .models import CashRegister, CashRegisterStatus, Invoice, InvoiceStatus, PaymentType


def _parse_date(value):
    """Parsea fecha desde el formulario (YYYY-MM-DD)."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _report_queryset(start, end):
    """Facturas emitidas y mermas en el rango (RF-12, UTC-5)."""
    qs = Invoice.objects.filter(status=InvoiceStatus.ISSUED)
    if start:
        qs = qs.filter(issued_at__date__gte=start)
    if end:
        qs = qs.filter(issued_at__date__lte=end)
    return qs


def _report_data(start, end):
    invoices = _report_queryset(start, end)
    gross_sales = invoices.aggregate(total=Sum("total"))["total"] or 0
    count = invoices.count()
    return invoices, gross_sales, count


def _guard_admin(request):
    if request.user.role != Role.ADMIN:
        messages.error(request, "No tiene permisos para acceder a este módulo.")
        return redirect(request.user.dashboard_url)
    return None


@login_required
def reports_view(request):
    """Consolidado de ventas brutas y mermas por rango de fechas (RF-12)."""
    blocked = _guard_admin(request)
    if blocked:
        return blocked

    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))

    invoices, gross_sales, count = _report_data(start, end)

    # Mermas en el rango (kitchen.Shrinkage vía OrderItem > Order.created_at)
    from kitchen.models import Shrinkage

    shrinkages = Shrinkage.objects.select_related("order_item", "registered_by")
    if start:
        shrinkages = shrinkages.filter(registered_at__date__gte=start)
    if end:
        shrinkages = shrinkages.filter(registered_at__date__lte=end)
    shrinkage_total = shrinkages.count()

    context = {
        "invoices": invoices,
        "gross_sales": gross_sales,
        "invoice_count": count,
        "shrinkages": shrinkages,
        "shrinkage_total": shrinkage_total,
        "start": start,
        "end": end,
    }
    return render(request, "billing/reports.html", context)


def _export_response(buffer, filename, content_type):
    """Descarga directa en el cliente: el archivo se genera en memoria (BytesIO)
    y se transmite como attachment sin escribir nada en el filesystem del servidor."""
    response = FileResponse(buffer, content_type=content_type, as_attachment=True, filename=filename)
    return response


@login_required
def reports_export_excel(request):
    """Exporta el consolidado a Excel .xlsx (RNF-09)."""
    blocked = _guard_admin(request)
    if blocked:
        return blocked

    from openpyxl import Workbook
    from openpyxl.styles import Font

    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    invoices, gross_sales, count = _report_data(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append(["Reporte de Ventas Brutas - Santy"])
    ws.append([f"Rango: {start or 'inicio'} a {end or 'hoy'} | Moneda: USD | Zona: UTC-5"])
    ws.append([])
    ws.append(["Factura #", "Fecha (UTC-5)", "Cliente", "Cédula", "Subtotal", "IVA", "Total", "Estado"])

    for row in ws.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.font = Font(bold=True)

    for inv in invoices:
        ws.append([
            inv.pk,
            inv.issued_at.strftime("%Y-%m-%d %H:%M") if inv.issued_at else "",
            inv.client_name,
            inv.client_cedula,
            float(inv.subtotal),
            float(inv.vat_amount),
            float(inv.total),
            inv.get_status_display(),
        ])

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", float(gross_sales)])

    # Hoja de mermas
    from kitchen.models import Shrinkage

    shrinkages = Shrinkage.objects.select_related("order_item", "registered_by")
    if start:
        shrinkages = shrinkages.filter(registered_at__date__gte=start)
    if end:
        shrinkages = shrinkages.filter(registered_at__date__lte=end)

    ws2 = wb.create_sheet("Mermas")
    ws2.append(["Reporte de Mermas - Santy"])
    ws2.append(["Motivo", "Platillo", "Registrado por", "Fecha (UTC-5)", "Reposición a $0.00"])
    for s in shrinkages:
        ws2.append([
            s.reason,
            s.order_item.name,
            str(s.registered_by),
            s.registered_at.strftime("%Y-%m-%d %H:%M"),
            "Sí" if s.notify_waiter_replacement else "No",
        ])

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return _export_response(
        buffer,
        f"reporte_santy_{start or 'all'}_{end or 'all'}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@login_required
def reports_export_pdf(request):
    """Exporta el consolidado a PDF (RNF-09)."""
    blocked = _guard_admin(request)
    if blocked:
        return blocked

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    invoices, gross_sales, count = _report_data(start, end)

    from kitchen.models import Shrinkage

    shrinkages = Shrinkage.objects.select_related("order_item", "registered_by")
    if start:
        shrinkages = shrinkages.filter(registered_at__date__gte=start)
    if end:
        shrinkages = shrinkages.filter(registered_at__date__lte=end)

    import io
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("Reporte de Ventas y Mermas — Santy", styles["Title"]))
    elements.append(Paragraph(
        f"Rango: {start or 'inicio'} a {end or 'hoy'} | Moneda USD | Zona UTC-5",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("Ventas brutas", styles["Heading2"]))
    data = [["Factura", "Fecha", "Cliente", "Cédula", "Subtotal", "IVA", "Total"]]
    for inv in invoices:
        data.append([
            str(inv.pk),
            inv.issued_at.strftime("%d/%m/%Y %H:%M") if inv.issued_at else "",
            inv.client_name,
            inv.client_cedula,
            f"{inv.subtotal:.2f}",
            f"{inv.vat_amount:.2f}",
            f"{inv.total:.2f}",
        ])
    data.append(["TOTAL", "", "", "", "", "", f"{gross_sales:.2f}"])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph("Mermas", styles["Heading2"]))
    m_data = [["Motivo", "Platillo", "Registrado por", "Fecha", "Reposición $0"]]
    for s in shrinkages:
        m_data.append([
            s.reason,
            s.order_item.name,
            str(s.registered_by),
            s.registered_at.strftime("%d/%m/%Y %H:%M"),
            "Sí" if s.notify_waiter_replacement else "No",
        ])
    m_table = Table(m_data, repeatRows=1)
    m_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    elements.append(m_table)

    doc.build(elements)
    buffer.seek(0)
    return _export_response(
        buffer,
        f"reporte_santy_{start or 'all'}_{end or 'all'}.pdf",
        "application/pdf",
    )


# ---------------------------------------------------------------------------
# Caja: apertura (RF-25/RF-26) y cierre ciego (RF-10/RF-11)
# ---------------------------------------------------------------------------


def _guard_cashier(request):
    if request.user.role != Role.CASHIER:
        messages.error(request, "No tiene permisos para acceder a este módulo.")
        return redirect(request.user.dashboard_url)
    return None


@login_required
def cash_register_open(request):
    """Abre la caja con fondo inicial > 0; rechaza fondo 0 (RF-25, RF-26)."""
    blocked = _guard_cashier(request)
    if blocked:
        return blocked

    if request.method == "POST":
        try:
            fund = float(request.POST.get("opening_fund", "").strip())
        except ValueError:
            messages.error(request, "Ingrese un fondo inicial válido.")
            return redirect("billing:cash_register_open")

        if fund <= 0:
            messages.error(
                request,
                "La apertura de caja requiere un fondo inicial mayor a $0.00. "
                "Las funciones de cobro permanecen bloqueadas.",
            )
            return redirect("billing:cash_register_open")

        # Solo una apertura activa por caja (invariante)
        if CashRegister.objects.filter(status=CashRegisterStatus.OPEN).exists():
            messages.error(request, "Ya existe una caja abierta en este turno.")
            return redirect("billing:cash_register_open")

        register = CashRegister.objects.create(
            opened_by=request.user,
            opening_fund=fund,
            status=CashRegisterStatus.OPEN,
        )
        messages.success(request, f"Caja abierta con fondo inicial de ${fund:.2f} (UTC-5).")
        return redirect("billing:cash_register_close")

    active = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
    return render(request, "billing/cash_register_open.html", {"active": active})


@login_required
def cash_register_close(request):
    """Cierre ciego: captura efectivo sin mostrar saldo esperado (RF-10/11)."""
    blocked = _guard_cashier(request)
    if blocked:
        return blocked

    register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()

    if request.method == "POST":
        if register is None:
            messages.error(request, "No hay caja abierta para cerrar.")
            return redirect("billing:cash_register_open")

        try:
            declared = float(request.POST.get("declared_cash", "").strip())
        except ValueError:
            messages.error(request, "Ingrese el efectivo declarado.")
            return redirect("billing:cash_register_close")

        register.close_blind(declared)

        if register.difference <= 2:
            register.closed_at = timezone.now()
            register.save(update_fields=["closed_at"])
            messages.success(request, f"Caja cuadrada. Diferencia: ${register.difference:.2f}")
            return redirect("billing:cash_register_open")

        # Descuadre: exige justificación para completar el cierre (RF-11)
        justification = request.POST.get("justification", "").strip()
        if not justification:
            messages.error(
                request,
                f"Descuadre de ${register.difference:.2f}. La justificación es obligatoria "
                "para completar el cierre.",
            )
            return render(request, "billing/cash_register_close.html", {"register": register})

        register.justify_and_close(justification)
        messages.success(request, "Cierre completado con descuadre justificado.")
        return redirect("billing:cash_register_open")

    if register is None:
        messages.info(request, "No hay caja abierta. Abra la caja para iniciar el turno.")
        return redirect("billing:cash_register_open")

    return render(request, "billing/cash_register_close.html", {"register": register})


# ---------------------------------------------------------------------------
# Facturación: emisión (RF-07), cobro parcial (RF-09), anulación (RF-08)
# ---------------------------------------------------------------------------


def _compute_totals(subtotal):
    """Aplica IVA configurable y redondea a 2 decimales."""
    from core.models import BusinessConfig

    config = BusinessConfig.objects.get(pk=1)
    subtotal = Decimal(str(subtotal)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    vat_amount = (subtotal * config.vat_rate).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return subtotal, vat_amount, subtotal + vat_amount


@login_required
def invoice_list(request):
    """Lista de facturas y cuentas listas para cobro (RF-07)."""
    blocked = _guard_cashier(request)
    if blocked:
        return blocked

    from django.utils import timezone
    from kitchen.models import Order, OrderStatus

    billable = (
        Order.objects.filter(status=OrderStatus.DELIVERED, invoice__isnull=True)
        .select_related("table", "waiter")
        .prefetch_related("items")
    )
    today = timezone.localdate()
    invoices = Invoice.objects.filter(issued_at__date=today).select_related("order__table")
    return render(
        request,
        "billing/invoice_list.html",
        {"billable": billable, "invoices": invoices},
    )


@login_required
def invoice_create(request, order_id):
    """Emite factura sobre una comanda: calcula IVA, cobra y libera la mesa."""
    blocked = _guard_cashier(request)
    if blocked:
        return blocked

    from django.db import transaction
    from django.utils import timezone
    from audit.models import ActionType, AuditLog, Result
    from kitchen.models import Order, OrderStatus
    from loyalty.models import LoyaltyMovement

    order = get_object_or_404(Order, pk=order_id)
    if order.status != OrderStatus.DELIVERED:
        messages.error(request, "La comanda debe estar entregada para facturar.")
        return redirect("billing:invoice_list")

    if hasattr(order, "invoice"):
        messages.error(request, "Esta comanda ya tiene factura.")
        return redirect("billing:invoice_list")

    from core.models import BusinessConfig

    config = BusinessConfig.objects.get(pk=1)
    items = order.items.all()
    subtotal = sum((i.quantity * i.unit_price for i in items), Decimal("0"))
    subtotal, vat_amount, total = _compute_totals(subtotal)

    context = {
        "order": order,
        "items": items,
        "subtotal": subtotal,
        "vat_amount": vat_amount,
        "total": total,
    }

    if request.method == "POST":
        client_name = request.POST.get("client_name", "").strip()
        client_cedula = request.POST.get("client_cedula", "").strip()
        paid_raw = request.POST.get("paid_amount", "").strip()
        points_raw = request.POST.get("redeem_points", "").strip()

        # Canje opcional de puntos (RF-16)
        discount = Decimal("0")
        if points_raw:
            try:
                points = int(points_raw)
                discount = Decimal(str(
                    LoyaltyMovement.redeem(client_cedula, points, float(subtotal))
                ))
                subtotal -= discount
                subtotal, vat_amount, total = _compute_totals(subtotal)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("billing:invoice_create", order_id=order.pk)

        try:
            paid = Decimal(str(float(paid_raw)))
        except (TypeError, ValueError):
            paid = total

        try:
            with transaction.atomic():
                register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
                invoice = Invoice.objects.create(
                    register=register,
                    order=order,
                    client_name=client_name,
                    client_cedula=client_cedula,
                    subtotal=subtotal,
                    vat_rate=config.vat_rate,
                    vat_amount=vat_amount,
                    total=total,
                    paid_amount=paid,
                    payment_type=PaymentType.PARTIAL if paid < total else PaymentType.FULL,
                    status=InvoiceStatus.DRAFT,
                )
                invoice.issue()

                if invoice.client_cedula:
                    LoyaltyMovement.accrue_for_invoice(invoice)
                if paid < total:
                    # RF-09: saldo pendiente asignado a la cédula
                    AuditLog.log(
                        request.user,
                        ActionType.INVOICE_ISSUE,
                        Result.SUCCESS,
                        object_type="Invoice",
                        object_id=invoice.pk,
                        detail=f"Saldo pendiente de {invoice.remaining_balance} USD a la cédula {client_cedula}.",
                    )
                AuditLog.log(
                    request.user,
                    ActionType.INVOICE_ISSUE,
                    Result.SUCCESS,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    detail=f"Factura #{invoice.pk} · total {invoice.total} USD · {invoice.get_payment_type_display()}.",
                )
        except Exception as exc:  # pragma: no cover
            messages.error(request, f"No fue posible emitir la factura: {exc}")
            return redirect("billing:invoice_list")

        messages.success(
            request,
            f"Factura #{invoice.pk} emitida · Total ${invoice.total:.2f} · "
            f"Pago ${invoice.paid_amount:.2f}.",
        )
        return redirect("billing:invoice_list")

    return render(request, "billing/invoice_create.html", context)


@login_required
def invoice_detail(request, invoice_id):
    """Detalle de una factura emitida (read-only)."""
    blocked = _guard_cashier(request)
    if blocked:
        return blocked
    invoice = get_object_or_404(Invoice, pk=invoice_id)
    return render(request, "billing/invoice_detail.html", {"invoice": invoice})


@login_required
def invoice_annul(request, invoice_id):
    """Anulación de factura Emitida: solo Administrador + PIN vigente (RF-08)."""
    if request.user.role != Role.ADMIN:
        messages.error(request, "No tiene permisos para acceder a este módulo.")
        return redirect(request.user.dashboard_url)

    from audit.models import ActionType, AuditLog, PinToken, Result

    invoice = get_object_or_404(Invoice, pk=invoice_id)

    if request.method == "POST":
        code = request.POST.get("pin", "").strip()
        token = PinToken.objects.filter(code=code).order_by("-issued_at").first()
        if token is None or not token.is_valid:
            AuditLog.log(
                request.user,
                ActionType.ANULATION,
                Result.FAILURE,
                object_type="Invoice",
                object_id=invoice.pk,
                detail="Intento de anulación sin PIN vigente.",
            )
            messages.error(request, "PIN inválido o expirado. La anulación fue denegada.")
            return redirect("billing:invoice_detail", invoice_id=invoice.pk)

        token.consumed_at = timezone.now()
        token.save(update_fields=["consumed_at"])
        invoice.annul(request.user)
        AuditLog.log(
            request.user,
            ActionType.ANULATION,
            Result.SUCCESS,
            object_type="Invoice",
            object_id=invoice.pk,
            detail=f"Factura #{invoice.pk} anulada ({invoice.total} USD).",
        )
        messages.success(request, f"Factura #{invoice.pk} anulada. Comprobante conservado.")
        return redirect("audit:trail")

    return render(request, "billing/invoice_annul.html", {"invoice": invoice})